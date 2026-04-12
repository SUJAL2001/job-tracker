# ============================================================
# excel_logger.py — Logs matched jobs to a formatted Excel sheet
# ============================================================

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

EXCEL_PATH = "job_tracker_results.xlsx"
SHEET_NAME = "Matched Jobs"

# ---- Colors ----
COLOR_HEADER_BG  = "2C7BE5"   # Blue header
COLOR_HEADER_FG  = "FFFFFF"   # White text
COLOR_HIGH       = "D4EDDA"   # Green bg  — score >= 85
COLOR_MED        = "FFF3CD"   # Yellow bg — score 70–84
COLOR_ALT_ROW    = "F8F9FA"   # Light grey for alternating rows
COLOR_BORDER     = "DEE2E6"

COLUMNS = [
    ("Date Found",       20),
    ("Company",          18),
    ("Job Title",        35),
    ("Match Score (%)",  16),
    ("AI Summary",       50),
    ("Skills Matched",   35),
    ("Skills Missing",   35),
    ("Job Link",         45),
    ("Status",           15),
]


def _thin_border():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _init_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"       # Freeze header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Job Tracker Summary"
    ws_sum["A1"].font = Font(bold=True, size=14, name="Arial")
    ws_sum["A3"] = "Total Matched Jobs"
    ws_sum["B3"] = f"=COUNTA('{SHEET_NAME}'!A2:A10000)"
    ws_sum["A4"] = "Avg Match Score"
    ws_sum["B4"] = f"=IFERROR(AVERAGE('{SHEET_NAME}'!D2:D10000),0)"
    ws_sum["A5"] = "High Matches (≥85%)"
    ws_sum["B5"] = f"=COUNTIF('{SHEET_NAME}'!D2:D10000,\">=85\")"
    ws_sum["A6"] = "Medium Matches (70–84%)"
    ws_sum["B6"] = f"=COUNTIFS('{SHEET_NAME}'!D2:D10000,\">=70\",'{SHEET_NAME}'!D2:D10000,\"<85\")"
    ws_sum["A8"] = "Last Updated"
    ws_sum["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    for row in range(3, 9):
        ws_sum.cell(row=row, column=1).font = Font(bold=True, name="Arial")
        ws_sum.cell(row=row, column=2).font = Font(name="Arial")
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 20

    return wb


def log_job_to_excel(job: dict, match_result: dict):
    """Append a matched job row to the Excel file. Creates file if needed."""
    score          = match_result.get("score", 0)
    reason         = match_result.get("reason", "")
    skills_matched = ", ".join(match_result.get("skills_matched", []))
    skills_missing = ", ".join(match_result.get("skills_missing", []))

    # Load or create workbook
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]
    else:
        wb = _init_workbook()
        ws = wb[SHEET_NAME]

    # Next empty row
    next_row = ws.max_row + 1

    row_data = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        job.get("company", ""),
        job.get("title", ""),
        score,
        reason,
        skills_matched,
        skills_missing,
        job.get("link", ""),
        "New",   # Status — user can change to Applied / Rejected / etc.
    ]

    # Pick row background color based on score
    if score >= 85:
        row_bg = COLOR_HIGH
    elif score >= 70:
        row_bg = COLOR_MED
    else:
        row_bg = COLOR_ALT_ROW if next_row % 2 == 0 else "FFFFFF"

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=row_bg)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = _thin_border()

        # Hyperlink for job link column
        if col_idx == 8 and value and value.startswith("http"):
            cell.hyperlink = value
            cell.font = Font(name="Arial", size=10, color="2C7BE5", underline="single")

        # Center-align score column
        if col_idx == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[next_row].height = 45

    # Update summary last-updated timestamp
    ws_sum = wb["Summary"]
    ws_sum["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb.save(EXCEL_PATH)
    print(f"[Excel] Logged: {job.get('title')} at {job.get('company')} (Score: {score}%)")


if __name__ == "__main__":
    # Quick test
    sample_job = {
        "title":   "Senior Python Developer",
        "company": "TestCorp",
        "link":    "https://example.com/job/123"
    }
    sample_result = {
        "score":          88,
        "reason":         "Strong Python and DevOps experience matches all core requirements.",
        "skills_matched": ["Python", "Docker", "CI/CD", "REST APIs"],
        "skills_missing": ["Kubernetes", "Go"]
    }
    log_job_to_excel(sample_job, sample_result)
    print(f"[Test] Excel file created: {EXCEL_PATH}")
